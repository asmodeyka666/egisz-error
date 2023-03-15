from tkinter import filedialog as fd
from tkinter import ttk
from tkinter import messagebox
import tkinter as tk
import pandas as pd
import os
import time
from datetime import datetime, timedelta
import numpy as np
from itertools import cycle, islice
from openpyxl import Workbook
import openpyxl as ox
from openpyxl.styles import (
                        PatternFill, Border, Side, 
                        Alignment, Font, GradientFill, numbers
                        )


SNILS_MR = {"МО": 'СНИЛС'}


def daily():
    file_names_zip = fd.askopenfilename(multiple=True, initialdir=r'C:\Users\User\Емиас\Корректировка ЕГИСЗ\выдача')
    if os.path.exists(os.path.dirname(file_names_zip[0]) + '\Раздача') == False:
        os.mkdir(os.path.dirname(file_names_zip[0]) + '\Раздача')
    print("Все файлы для обработки:", file_names_zip, sep='/n')
    for file_name in file_names_zip:
        print("Рассматриваем: ", os.path.basename(file_name))
        if '2196' in file_name:
            modify(file_name)
        elif '2244' in file_name:
            modify_one(file_name)
        elif 'НеИД' in file_name or '2389' in file_name:
            modify_noID(file_name)
    root.destroy()
        
# Разбивка Идов
def modify(file_name):
    if file_name == '':
        file_name = fd.askopenfilename()

    # забираем файл
    xls = pd.ExcelFile(file_name)

    list_send_1 = ''
    list_send_2 = ''
    for lists_xls in xls.sheet_names:
        if '1' in lists_xls:
            if 'пере' in lists_xls or 'от' in lists_xls:
                list_send_1 = lists_xls
            else:
                list_1 = lists_xls
        elif '2' in lists_xls:
            if 'пере' in lists_xls or 'от' in lists_xls:
                list_send_2 = lists_xls
            else:
                list_2 = lists_xls
            
    df1 = pd.read_excel(file_name, list_1, dtype=str)
    df2 = pd.read_excel(file_name, list_2, dtype=str)
    if list_send_1 != '':
        df1_send_1 = pd.read_excel(file_name, list_send_1, dtype=str)
    if list_send_2 != '':
        df1_send_2 = pd.read_excel(file_name, list_send_2, dtype=str)
    df1 = pd.concat([df1, df2], ignore_index=True)
    df1 = df1.fillna('')
    df1.insert(0, 'Внесено', '')
    df1.insert(1, 'Комментарии', '')
    # не внесено
    df_not_included = df1.sort_values(['Описание ошибки', 'Наменование МО', 'Фамилия пациента', 'Имя пациента', 'Отчество', 'Дата рождения'],
                                           ascending=[True, True, True, True, True, True])
    df1_not_included = df_not_included.loc[(df_not_included['Кратность вакцинации'] == 'V1') | (df_not_included['Кратность вакцинации'] == 'переотправка V1')]
    df2_not_included = df_not_included.loc[(df_not_included['Кратность вакцинации'] == 'V2') | (df_not_included['Кратность вакцинации'] == 'переотправка V2')]        
    all_sheets = {'V1': df1_not_included, 'V2': df2_not_included}

    writer = pd.ExcelWriter(os.path.dirname(file_name) + r'\не внесено_' + os.path.basename(file_name),
                                engine='xlsxwriter')

    for sheet_name in all_sheets.keys():
        all_sheets[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)

    writer.save()

    df1['СНИЛС МР верный'] = ''
    df1 = df1.fillna('')
    sortdf = ['Внесено', 'Комментарии', 'ИД пациента', 'Описание ошибки', 'Кратность вакцинации', 'СНИСЛ (из ЕРП)',
              'СНИЛС пациента из документа',
              'Фамилия пациента', 'Имя пациента', 'Отчество', 'Пол', 'Дата рождения', 'Наименование типа ДУЛ',
              'Серия ДУЛ',
              'Номер ДУЛ', 'Дата выдачи', 'Полис ОМС', 'Контакный телефон', 'Мобильный телефон',
              'Адрес_регистрации_город',
              'Адрес_регистрации_улица', 'Адрес_регистрации_дом', 'Адрес_проживания_город', 'Адрес_проживания_улица',
              'Адрес_проживания_дом', 'Дата вакцинации оцифровка', 'Наменование МО', 'СНИЛС МР',
              'Температура тела', 'Препарат вакцины', 'GTIN', 'Серийный номер (ISN)', 'Серия и контрольный номер',
              'Дата вакцинации', 'Статус передачи', 'Код типа ДУЛ', 'Наименование МУ', 
              'Допуск к вакцинации', 'Производитель', 'Срок годности', 'Жалобы на момент осмотра', 'Место введения',
              'Наличие противопоказаний', 'Побочная реакциия на прививку', 'ФИО МР', 'СНИЛС МР верный', 'DOCUMENT_ID', 'DOCUMENT_CREATED',
              'CCT', 'Адрес_регистрации_корпус', 'Адрес_регистрации_строение', 'Адрес_регистрации_квартира',
              'Адрес_регистрации_строкой', 'Адрес_регистрации_код_кладр', 'Адрес_проживания_корпус',
              'Адрес_проживания_строение', 'Адрес_проживания_квартира', 'Адрес_проживания_строкой',
              'Адрес_проживания_код_КЛАДР']

    #Заменяем СНИЛС мед работника на значение из базы
    #df1['СНИЛС МР верный'] = df1['Наменование МО'].map(SNILS_MR)
    #df1['СНИЛС МР верный'].fillna(df1['СНИЛС МР'], inplace=True)
    
    df1 = df1[sortdf]
    df1['GTIN'] = df1['GTIN'].str.replace('046', '')
    df1['Пол'] = df1['Пол'].str.replace('1', 'M')
    df1['Пол'] = df1['Пол'].str.replace('2', 'Ж')

    df1['Дата вакцинации оцифровка'] = df1['Дата вакцинации оцифровка'].combine(df1['Дата вакцинации'],
                                                                                (lambda x1, x2: x1 if x1 != '' else x2))
    
    df1_filtr = df1.sort_values(['Кратность вакцинации', 'Описание ошибки', 'Наменование МО', 'Фамилия пациента', 'Имя пациента', 'Отчество', 'Дата рождения'],
                                           ascending=[True, True, True, True, True, True, True])

    df1_filtr['Дата рождения'] = df1_filtr['Дата рождения'].map(lambda x: x[0: x.find('T')] if 'T' in x else x)
    df1_filtr['Дата выдачи'] = df1_filtr['Дата выдачи'].map(lambda x: x[0: x.find('T')] if 'T' in x else x)
    df1_filtr['Дата вакцинации оцифровка'] = df1_filtr['Дата вакцинации оцифровка'].map(lambda x: x[0: x.find('T')] if 'T' in x else x)
    

    df1_filtr['Дата рождения'] = (pd.to_datetime(df1_filtr['Дата рождения'], errors='coerce')
                          .dt.strftime("%d.%m.%Y"))
    
    df1_filtr['Дата выдачи'] = (pd.to_datetime(df1_filtr['Дата выдачи'], errors='coerce')
                          .dt.strftime("%d.%m.%Y"))
    
    df1_filtr['Дата вакцинации оцифровка'] = (pd.to_datetime(df1_filtr['Дата вакцинации оцифровка'], errors='coerce')
                          .dt.strftime("%d.%m.%Y"))
    
    
    df1_filtr = df1_filtr.iloc[:, :33]

    #Проверяем на различие СНИЛСов пациентов (реализовано ниже заливкой ячейки в openpyxl)
    #df1_filtr.insert(7, 'Сравнение СНИЛС', '')
    #df1_filtr['Сравнение СНИЛС'] = ((df1_filtr.loc[:, 'СНИСЛ (из ЕРП)'] == df1_filtr.loc[:, 'СНИЛС пациента из документа']) |
                                    #(df1_filtr['СНИСЛ (из ЕРП)'] == '') | (df1_filtr['СНИЛС пациента из документа'] == ''))
    
        
    df_give = df1_filtr.loc[df1_filtr['Кратность вакцинации'].isin(['V1', 'V2'])]
    df_give.to_excel(os.path.dirname(file_name) + '\Раздача' + r'\Раздача_' + os.path.basename(file_name), index=False)
    
    df1_filtr_not_included = df1_filtr.loc[df1_filtr['Кратность вакцинации'].isin(['переотправка V1', 'переотправка V2'])]
    df1_filtr_not_included = df1_filtr_not_included.sort_values(['Кратность вакцинации', 'Фамилия пациента', 'Имя пациента', 'Отчество', 'Дата рождения'],
                                           ascending=[True, True, True, True, True])
    if len(df1_filtr_not_included) > 0:
        df1_filtr_not_included.to_excel(os.path.dirname(file_name) + r'\Переотправка_' + os.path.basename(file_name), index=False)
    #форматируем файл раздача
    file_give = os.path.dirname(file_name) + '\Раздача' + r'\Раздача_' + os.path.basename(file_name)
    wb = ox.load_workbook(filename=file_give, read_only=False)
    sheet = wb.active
    # Получение даты из названия файла
    file_date = os.path.basename(file_give).split('.')[0].split('_')[-1]
    file_date = datetime.strptime(file_date, "%Y-%m-%d").date()
    # Цикл по ячейкам столбца Дата
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 26)
        cell_date = datetime.strptime(cell.value, "%d.%m.%Y").date()
        # Расчет разницы между ячейкой и указанной датой
        delta = abs((file_date - cell_date).days)
        # Проверка разницы на меньше или равно 1 дню
        if delta > 1:
            cell.fill = ox.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", patternType="solid")
    for row in range(2, sheet.max_row + 1):
        cell1 = sheet.cell(row, 6)
        cell2 = sheet.cell(row, 7)
        flag = (cell1.value == cell2.value or cell1.value == None or cell2.value == None)
        if flag == False:
            cell1.fill = ox.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", patternType="solid")
            cell2.fill = ox.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", patternType="solid")

    sheet.column_dimensions['A'].width = 3.6
    sheet.column_dimensions['B'].width = 3.6
    sheet.column_dimensions['E'].width = 3
    sheet.column_dimensions['K'].width = 2.5
    sheet.column_dimensions['L'].width = 10
    sheet.column_dimensions['N'].width = 4.3
    sheet.column_dimensions['O'].width = 5.8
    sheet.column_dimensions['T'].width = 7
    sheet.column_dimensions['T'].width = 7
    sheet.column_dimensions['V'].width = 3
    sheet.column_dimensions['W'].width = 7
    sheet.column_dimensions['Y'].width = 3
    sheet.column_dimensions['Z'].width = 10
    sheet.column_dimensions['AC'].width = 3
        
    
    wb.save(file_give)
    wb.close()

    os.startfile(os.path.dirname(file_give))



def modify_one(file_name):
    if file_name == '':
        file_name = fd.askopenfilename()

    # забираем файл
    df1 = pd.read_excel(file_name, dtype=str)
    
    df1 = df1.fillna('')
    df1['СНИЛС МР верный'] = ''
    sortdf = ['ID_EMIAS', 'ERROR_DESCRIPTION', 'Кратность вакцинации', 'СНИСЛ (из ЕРП)',
              'СНИЛС пациента из документа',
              'Фамилия пациента', 'Имя пациента', 'Отчество', 'Пол', 'Дата рождения', 'Наименование типа ДУЛ',
              'Серия ДУЛ',
              'Номер ДУЛ', 'Дата выдачи', 'Полис ОМС', 'Контакный телефон', 'Мобильный телефон',
              'Адрес_регистрации_город',
              'Адрес_регистрации_улица', 'Адрес_регистрации_дом', 'Адрес_проживания_город', 'Адрес_проживания_улица',
              'Адрес_проживания_дом', 'Дата вакцинации оцифровка', 'Наменование МО', 'СНИЛС МР',
              'Температура тела', 'Препарат вакцины', 'GTIN', 'Серийный номер (ISN)', 'Серия и контрольный номер',
              'Дата вакцинации', 'STATUS_ERROR', 'Код типа ДУЛ', 'Наименование МУ', 
              'Допуск к вакцинации', 'Производитель', 'Срок годности', 'Жалобы на момент осмотра', 'Место введения',
              'Наличие противопоказаний', 'Побочная реакциия на прививку', 'ФИО МР', 'СНИЛС МР верный', 'DOCUMENT_ID', 'DOCUMENT_CREATED',
              'CCT', 'Адрес_регистрации_корпус', 'Адрес_регистрации_строение', 'Адрес_регистрации_квартира',
              'Адрес_регистрации_строкой', 'Адрес_регистрации_код_кладр', 'Адрес_проживания_корпус',
              'Адрес_проживания_строение', 'Адрес_проживания_квартира', 'Адрес_проживания_строкой',
              'Адрес_проживания_код_КЛАДР']

    #Заменяем СНИЛС мед работника на значение из базы
    #df1['СНИЛС МР верный'] = df1['Наменование МО'].map(SNILS_MR)
    #df1['СНИЛС МР верный'].fillna(df1['СНИЛС МР'], inplace=True)


    df1 = df1[sortdf]
    df1['GTIN'] = df1['GTIN'].str.replace('046', '')
    df1['Пол'] = df1['Пол'].str.replace('1', 'M')
    df1['Пол'] = df1['Пол'].str.replace('2', 'Ж')

    df1['Дата вакцинации оцифровка'] = df1['Дата вакцинации оцифровка'].combine(df1['Дата вакцинации'],
                                                                                (lambda x1, x2: x1 if x1 != '' else x2))
    
    df1_filtr = df1.sort_values(['Фамилия пациента', 'Имя пациента', 'Дата рождения'],
                                           ascending=[True, True, True])
    
    df1_filtr['Дата рождения'] = df1_filtr['Дата рождения'].map(lambda x: x[0: x.find('T')] if 'T' in x else x)
    df1_filtr['Дата выдачи'] = df1_filtr['Дата выдачи'].map(lambda x: x[0: x.find('T')] if 'T' in x else x)
    df1_filtr['Дата вакцинации оцифровка'] = df1_filtr['Дата вакцинации оцифровка'].map(lambda x: x[0: x.find('T')] if 'T' in x else x)
    
    df1_filtr['Дата рождения'] = (pd.to_datetime(df1_filtr['Дата рождения'], errors='coerce')
                          .dt.strftime("%d.%m.%Y"))
    df1_filtr['Дата выдачи'] = (pd.to_datetime(df1_filtr['Дата выдачи'], errors='coerce')
                          .dt.strftime("%d.%m.%Y"))
    df1_filtr['Дата вакцинации оцифровка'] = (pd.to_datetime(df1_filtr['Дата вакцинации оцифровка'], errors='coerce')
                          .dt.strftime("%d.%m.%Y"))
    
    df1_filtr = df1_filtr.iloc[:, :31]

    df1_filtr.to_excel(os.path.dirname(file_name) + '\Раздача' + r'\Пересобранный_' + os.path.basename(file_name), index=False)





def modify_noID(file_name):
    if file_name == '':
        file_name = fd.askopenfilename()

    if 'v1_по_v2' not in file_name and 'v1' not in file_name and 'v2' not in file_name and 'V1' not in file_name and 'V2' not in file_name and 'в1' not in file_name and 'В1' not in file_name and 'в2' not in file_name and 'В2' not in file_name:
    # забираем файл
        xls = pd.ExcelFile(file_name)

        list_1 = ''
        list_2 = ''
        for lists_xls in xls.sheet_names:
            if '1' in lists_xls:
                if 'пере' in lists_xls or 'от' in lists_xls:
                    list_send_1 = lists_xls
                else:
                    list_1 = lists_xls
            elif '2' in lists_xls:
                if 'пере' in lists_xls or 'от' in lists_xls:
                    list_send_2 = lists_xls
                else:
                    list_2 = lists_xls
            
        if list_1 != '':
            df1 = pd.read_excel(file_name, list_1, dtype=str)
        else:
            df1 = pd.DataFrame()
        if list_2 != '':
            df2 = pd.read_excel(file_name, list_2, dtype=str)
        else:
            df2 = pd.DataFrame()
        df1 = pd.concat([df1, df2], ignore_index=True)
        df1 = df1.fillna('')

        df1.insert(0, 'Внесено', '')
        df1.insert(1, 'Комментарии', '')
        # не внесено
        df_not_included = df1.sort_values(['Кратность вакцинации', 'Описание ошибки', 'Наменование МО', 'Фамилия', 'Имя', 'Отчество', 'Дата рождения'],
                                           ascending=[True, True, True, True, True, True, True])
        df1_not_included = df_not_included.loc[(df_not_included['Кратность вакцинации'] == 'V1') | (df_not_included['Кратность вакцинации'] == 'переотправка V1')]
        df2_not_included = df_not_included.loc[(df_not_included['Кратность вакцинации'] == 'V2') | (df_not_included['Кратность вакцинации'] == 'переотправка V2')]        
        all_sheets = {'V1' : df1_not_included, 'V2' : df2_not_included}

        writer = pd.ExcelWriter(os.path.dirname(file_name) + r'\не внесено_' + os.path.basename(file_name),
                                engine='xlsxwriter')

        for sheet_name in all_sheets.keys():
            all_sheets[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)

        writer.save()

        df1['СНИЛС МР верный'] = ''
        df1 = df1.fillna('')
        sortdf = ['Внесено', 'Комментарии', 'UNIDENTIFIED_PATIENT_ID', 'Описание ошибки', 'Кратность вакцинации', 'СНИЛС',
              'Фамилия', 'Имя', 'Отчество', 'Дата рождения', 'Наименование ДУЛ',
              'Серия дул', 'Номер ДУЛ', 'Дата выдачи ДУЛ', 'Наименование документа иностарнного гаржданина', 'Серия документа иностранного гражданина',
              'Номер документа иностранного гражданина', 'Полис ОМС', 'Город',
              'Улица', 'Дом', 'Дата вакцинации', 'Наменование МО', 'СНИЛС МР',
              'Препарат вакцины', 'GTIN', 'Серийный номер (ISN)', 'Серия и контрольный номер',
              'DOCUMENT_CREATED', 'CCT', 'Признак иностранного гражданина', 'Код типа ДУЛ', 'Код типа документа иностранного гражданина',
              'Регион', 'Наименование МУ', 'Допуск к вакцинации', 'ФИО МР', 'СНИЛС МР верный', 'Дата вакцинации', 'Производитель']
        #Заменяем СНИЛС мед работника на значение из базы
        #df1['СНИЛС МР верный'] = df1['Наменование МО'].map(SNILS_MR)
        #df1['СНИЛС МР верный'].fillna(df1['СНИЛС МР'], inplace=True)

        df1 = df1[sortdf]
        df1['GTIN'] = df1['GTIN'].str.replace('046', '')

        df1_filtr = df1.sort_values(['Кратность вакцинации', 'Описание ошибки', 'Наменование МО', 'Фамилия', 'Имя', 'Отчество', 'Дата рождения'],
                                           ascending=[True, True, True, True, True, True, True])
        
        df1_filtr = df1_filtr.iloc[:, :-12]
        
        df1_filtr['Дата рождения'] = df1_filtr['Дата рождения'].map(lambda x: x[0: x.find('T')] if 'T' in x else x)
        df1_filtr['Дата выдачи ДУЛ'] = df1_filtr['Дата выдачи ДУЛ'].map(lambda x: x[0: x.find('T')] if 'T' in x else x)
        df1_filtr['Дата вакцинации'] = df1_filtr['Дата вакцинации'].map(lambda x: x[0: x.find('T')] if 'T' in x else x)
        

        df1_filtr['Дата рождения'] = (pd.to_datetime(df1_filtr['Дата рождения'], errors='coerce')
                          .dt.strftime("%d.%m.%Y"))
        
        df1_filtr['Дата выдачи ДУЛ'] = (pd.to_datetime(df1_filtr['Дата выдачи ДУЛ'], errors='coerce')
                          .dt.strftime("%d.%m.%Y"))
        
        df1_filtr['Дата вакцинации'] = (pd.to_datetime(df1_filtr['Дата вакцинации'], errors='coerce')
                          .dt.strftime("%d.%m.%Y"))

        df1_filtr.loc[df1_filtr['Кратность вакцинации'].isin(['V1', 'V2'])].to_excel(os.path.dirname(file_name) + '\Раздача' + r'\Раздача_' + os.path.basename(file_name), index=False)
        df1_filtr_not_included = df1_filtr.loc[df1_filtr['Кратность вакцинации'].isin(['переотправка V1', 'переотправка V2'])]
        df1_filtr_not_included = df1_filtr_not_included.sort_values(['Кратность вакцинации', 'Фамилия', 'Имя', 'Отчество', 'Дата рождения'],
                                           ascending=[True, True, True, True, True])
        if len(df1_filtr_not_included) > 0:
            df1_filtr_not_included.to_excel(os.path.dirname(file_name) + r'\Переотправка_' + os.path.basename(file_name), index=False)
              

        

    else:
        df1 = pd.read_excel(file_name, dtype=str)
        df1 = df1.fillna('')
        df1['СНИЛС МР верный'] = ''
        sortdf = ['UNIDENTIFIED_PATIENT_ID', 'Описание ошибки', 'Кратность вакцинации', 'СНИЛС',
              'Фамилия', 'Имя', 'Отчество', 'Дата рождения', 'Наименование ДУЛ',
              'Серия дул', 'Номер ДУЛ', 'Дата выдачи ДУЛ', 'Наименование документа иностарнного гаржданина', 'Серия документа иностранного гражданина',
              'Номер документа иностранного гражданина', 'Полис ОМС', 'Город',
              'Улица', 'Дом', 'Дата вакцинации', 'Наменование МО', 'СНИЛС МР',
              'Препарат вакцины', 'GTIN', 'Серийный номер (ISN)', 'Серия и контрольный номер',
              'DOCUMENT_CREATED', 'CCT', 'Признак иностранного гражданина', 'Код типа ДУЛ', 'Код типа документа иностранного гражданина',
              'Регион', 'Наименование МУ', 'Допуск к вакцинации', 'ФИО МР', 'СНИЛС МР верный', 'Дата вакцинации', 'Производитель']
        #Заменяем СНИЛС мед работника на значение из базы
        #df1['СНИЛС МР верный'] = df1['Наменование МО'].map(SNILS_MR)
        #df1['СНИЛС МР верный'].fillna(df1['СНИЛС МР'], inplace=True)

        df1 = df1[sortdf]
        df1['GTIN'] = df1['GTIN'].str.replace('046', '')

        df1_filtr = df1.sort_values(['Фамилия', 'Имя', 'Дата рождения'],
                                           ascending=[True, True, True])

        df1_filtr = df1_filtr.iloc[:, :26]
        
        df1_filtr['Дата рождения'] = df1_filtr['Дата рождения'].map(lambda x: x[0: x.find('T')] if 'T' in x else x)
        df1_filtr['Дата рождения'] = (pd.to_datetime(df1_filtr['Дата рождения'], errors='coerce', format = '%Y-%m-%d')
                          .dt.strftime("%d.%m.%Y"))
        df1_filtr['Дата выдачи ДУЛ'] = df1_filtr['Дата выдачи ДУЛ'].map(lambda x: x[0: x.find('T')] if 'T' in x else x)
        df1_filtr['Дата выдачи ДУЛ'] = (pd.to_datetime(df1_filtr['Дата выдачи ДУЛ'], errors='coerce', format = '%Y-%m-%d')
                          .dt.strftime("%d.%m.%Y"))
        df1_filtr['Дата вакцинации'] = df1_filtr['Дата вакцинации'].map(lambda x: x[0: x.find('T')] if 'T' in x else x)
        df1_filtr['Дата вакцинации'] = (pd.to_datetime(df1_filtr['Дата вакцинации'], errors='coerce', format = '%Y-%m-%d')
                          .dt.strftime("%d.%m.%Y"))

        df1_filtr.to_excel(os.path.dirname(file_name) + '\Раздача' + r'\Пересобранный_' + os.path.basename(file_name), index=False)


def totals():

    def data_file(file_name):
    # Получение даты из названия файла
        file_date = file_name.split('.')[0].split('_')[-1]
        file_date = datetime.strptime(file_date, "%Y-%m-%d").date()
        return file_date

        
    def no_enter(file_name, df_give):
        print('Добавление в файл "Не внесено"')
        file_name = file_name.split('Раздача_')[1]
        wb = ox.load_workbook(filename=folder+'\\не внесено_'+file_name, read_only=False)
        for sheet_name in wb.sheetnames:
            df_give_res = df_give.loc[df_give['Кратность вакцинации'].isin([sheet_name])].iloc[:,:2]
            df_give_res['Внесено'] = df_give_res['Внесено'].astype(int)
            ws = wb[sheet_name]
            for ir in range(0, len(df_give_res)):
                for ic in range(0, len(df_give_res.iloc[ir])):
                    ws.cell(row=2 + ir,column=1 + ic).value = df_give_res.iloc[ir][ic]
       
            wb.save(folder+'\\не внесено_'+file_name)
        wb.close()


    def no_enter_all(file_name):
        print('Добавление в файл "Не внесено общий"')
        file_name = file_name.split('Раздача_')[1]
        df_no_entry1 = pd.read_excel(folder+'\\не внесено_'+file_name, 'V1', dtype=str)
        df_no_entry2 = pd.read_excel(folder+'\\не внесено_'+file_name, 'V2', dtype=str)
        df_no_entry = pd.concat([df_no_entry1, df_no_entry2], ignore_index=True)
        df_no_entry = df_no_entry.loc[df_no_entry['Внесено'].isin(['0'])]
        df_no_entry['Внесено'] = df_no_entry['Внесено'].astype(int)

        #Проверка наличие базы
        file_no_enter_all = r'D:\Мои файлы\Documents\Емиас\Корректировка ЕГИСЗ\Не внесено на 28.10.22.xlsx'
        if os.path.isfile(file_no_enter_all):
            print ('Файл "Не внесено ИД" найден')
        else:
            print ('Файл с базой НЕ найден!!!')
            print ('Выберите файл "Не внесено ИД" ')
            messagebox.showinfo('Выберите файл', 'Файл "Не внесено ИД" НЕ найден!!')
            file_no_enter_all = fd.askopenfilename(title='Выберите файл "Не внесено ИД"', initialdir=r'C:\Users\User\Емиас\Корректировка ЕГИСЗ')

        wb = ox.load_workbook(filename=file_no_enter_all, read_only=False)
        ws = wb['Не внесено']
        # Выбор номера строки для начала добавления данных
        start_row = ws.max_row + 1
        #Заполнение датой
        for ir in range(0, len(df_no_entry)):
            ws.cell(row=start_row + ir, column=1).value = data_file(file_name)
            ws.cell(row=start_row + ir, column=1).number_format = 'DD.MM.YYYY'
        #Заполнение данными из df
        for ir in range(0, len(df_no_entry)):
            for ic in range(0, len(df_no_entry.iloc[ir])):
                ws.cell(row=start_row + ir,column=4 + ic).value = df_no_entry.iloc[ir][ic]
            
       
        wb.save(file_no_enter_all)
        wb.close()
        

    def send_MO(file_name, df_give):
        print('Добавление в файл "разбор МО"')
        df_send_MO = df_give.loc[df_give['Внесено'].str.startswith('0') | (df_give['Комментарии'].notnull()
                                                                           & ~df_give['Комментарии'].astype(str).str.lower().str.startswith('изменить')
                                                                           & ~df_give['Комментарии'].astype(str).str.lower().str.startswith('удали'))]
        df_send_MO.loc[:, 'Внесено'] = df_send_MO.loc[:, 'Внесено'].astype(int)

        #Проверка наличие базы
        file_send_MO = r'D:\Мои файлы\Documents\Емиас\Корректировка ЕГИСЗ\разбор МО от 31.10.22.xlsx'
        if os.path.isfile(file_send_MO):
            print ('Файл "Разбор в МО" найден')
        else:
            messagebox.showinfo('Выберите файл', 'Файл "Разбор в МО" НЕ найден!!')
            file_send_MO = fd.askopenfilename(title='Выберите файл "Разбор в МО"', initialdir=r'C:\Users\User\Емиас\Корректировка ЕГИСЗ')
                
        if '2196' in file_name:
            sheet_name = 'Иды'    
        elif 'НеИД' in file_name or '2389' in file_name:
            sheet_name = 'неиды'
            
        wb = ox.load_workbook(filename=file_send_MO, read_only=False)
        ws = wb[sheet_name]
        # Выбор номера строки для начала добавления данных
        start_row = ws.max_row + 1
        #Заполнение данными из df
        for ir in range(0, len(df_send_MO)):
            for ic in range(0, len(df_send_MO.iloc[ir])):
                ws.cell(row=start_row + ir,column=2 + ic).value = df_send_MO.iloc[ir][ic]
            
        wb.save(file_send_MO)
        wb.close()   

        os.startfile(file_send_MO)
        
    def report(file_name, df_give):
        #Проверка наличие файла Отчет
        file_report = r'D:\Мои файлы\Documents\Емиас\Корректировка ЕГИСЗ\Ежедневный google отчет.xlsx'
        if os.path.isfile(file_report):
            print ('Файл "Ежедневный google отчет" найден')
        else:
            messagebox.showinfo('Выберите файл', 'Файл "Ежедневный google отчет" НЕ найден!!')
            file_report = fd.askopenfilename(title='Выберите файл "Ежедневный google отчет"', initialdir=r'C:\Users\User\Емиас\Корректировка ЕГИСЗ')
        wb = ox.load_workbook(filename=file_report, read_only=False)
        ws = wb['Вакцинация']
        start_row = ws.max_row + 1
        #ищем первую пустую строку столбца А
        for ir in range(800, ws.max_row + 1):
            if ws.cell(row=ir, column=1).value == None:
                start_row = ir
                break
                
        if '2196' in file_name:
            ws.cell(row=start_row,column=1).value = data_file(file_name).strftime('%d.%m.%Y')
            ws.cell(row=start_row,column=2).value = len(df_give.loc[df_give['Кратность вакцинации'].str.startswith('V1')])
            ws.cell(row=start_row,column=3).value = len(df_give.loc[df_give['Комментарии'].isin(['ЧМО']) & df_give['Кратность вакцинации'].str.startswith('V1')])
            ws.cell(row=start_row,column=4).value = f'=B{start_row}-C{start_row}'
            ws.cell(row=start_row,column=5).value = len(df_give.loc[df_give['Внесено'].isin(['1']) & df_give['Кратность вакцинации'].str.startswith('V1')])
            ws.cell(row=start_row,column=6).value = len(df_give.loc[df_give['Внесено'].isin(['0']) & df_give['Кратность вакцинации'].str.startswith('V1')])
            if ws.cell(row=start_row,column=6).value != 0:
                ws.cell(row=start_row,column=7).value = 'на разбор в МО - ' + str(len(df_give.loc[df_give['Внесено'].isin(['0']) & df_give['Кратность вакцинации'].str.startswith('V1')]))
            ws.cell(row=start_row,column=8).value = len(df_give.loc[df_give['Внесено'].isin(['2']) & df_give['Кратность вакцинации'].str.startswith('V1')])
            ws.cell(row=start_row,column=9).value = f'=C{start_row}+E{start_row}+F{start_row}+H{start_row}'
            ws.cell(row=start_row,column=10).value = len(df_give.loc[df_give['Кратность вакцинации'].str.startswith('V2')])
            ws.cell(row=start_row,column=11).value = len(df_give.loc[df_give['Комментарии'].isin(['ЧМО']) & df_give['Кратность вакцинации'].str.startswith('V2')])
            ws.cell(row=start_row,column=12).value = f'=J{start_row}-K{start_row}'
            ws.cell(row=start_row,column=13).value = len(df_give.loc[df_give['Внесено'].isin(['1']) & df_give['Кратность вакцинации'].str.startswith('V2')])
            ws.cell(row=start_row,column=14).value = len(df_give.loc[df_give['Внесено'].isin(['0']) & df_give['Кратность вакцинации'].str.startswith('V2')])
            if ws.cell(row=start_row,column=14).value != 0:
                ws.cell(row=start_row,column=15).value = 'на разбор в МО - ' + str(len(df_give.loc[df_give['Внесено'].isin(['0']) & df_give['Кратность вакцинации'].str.startswith('V2')]))
            ws.cell(row=start_row,column=16).value = len(df_give.loc[df_give['Внесено'].isin(['2']) & df_give['Кратность вакцинации'].str.startswith('V2')])
            ws.cell(row=start_row,column=17).value = f'=K{start_row}+M{start_row}+N{start_row}+P{start_row}'
            ws.cell(row=start_row,column=18).value = f'=B{start_row}+J{start_row}'
            ws.cell(row=start_row,column=19).value = f'=E{start_row}+M{start_row}+P{start_row}+H{start_row}'
            ws.cell(row=start_row,column=20).value = f'=S{start_row}/(B{start_row}+J{start_row})'
            ws.cell(row=start_row,column=22).value = f'=R{start_row}-S{start_row}-K{start_row}-C{start_row}'
            ws.cell(row=start_row,column=20).number_format = '0.00%'
            
            
        elif 'НеИД' in file_name or '2389' in file_name:
            pass
        ws.cell(row=start_row+1,column=1).value = data_file(file_name).strftime('%d.%m.%Y') + '\nНЕИД'
        
        wb.save(file_report)
        
        #внесени записей для изменения в ЕМИАС
        ws = wb['Редактирование в ЕМИАС']
        df_edit = df_give.loc[df_give['Комментарии'].astype(str).str.lower().str.startswith('измен') | df_give['Комментарии'].astype(str).str.lower().str.startswith('удали')]
        df_edit = df_edit.loc[:, (['ИД пациента', 'Полис ОМС', 'Фамилия пациента', 'Имя пациента', 'Отчество', 'Дата рождения', 'Комментарии'])]

        #ищем первую пустую строку
        start_row = ws.max_row + 1


            
        for ir in range(0, len(df_edit)):
            ws.cell(row=start_row + ir, column=9).value = data_file(file_name)
            ws.cell(row=start_row + ir, column=9).number_format = 'DD.MM.YYYY'
            ws.cell(row=start_row + ir, column=1).value = start_row + ir + 1
            for ic in range(0, len(df_edit.iloc[ir])):
                ws.cell(row=start_row + ir, column=2 + ic).value = df_edit.iloc[ir][ic]

        
        
        wb.save(file_report)
        wb.close()
        
        os.startfile(file_report)
        
    
    #читаем файлы раздачи
    global folder
    folder = fd.askdirectory(title='Выберите папку за день')
    for file_name in os.listdir(folder + r'\\Раздача\\'):
        if file_name.startswith('Раздача_'):
            if '2196' in file_name and 'Женя' not in file_name:
                df_give = pd.read_excel(folder + '\\Раздача\\' + file_name, dtype=str)
                no_enter(file_name, df_give)
                no_enter_all(file_name)
                send_MO(file_name, df_give)
                report(file_name, df_give)
            elif 'НеИД' in file_name or '2389' in file_name:
                df_give_noid = pd.read_excel(folder + '\\Раздача\\' + file_name, dtype=str)

    root.destroy()

            

    
root = tk.Tk()
root.title("Обработка запросов")
root.geometry('400x200')
root["bg"] = "#fff"


button1 = tk.Button(text="Ежедневная модификация",
                   command=daily, background="#fff", foreground="#3b3e41",
                   padx="30", pady="15", font="15")

button2 = tk.Button(text="Распределение итогов",
                   command=totals, background="#fff", foreground="#3b3e41",
                   padx="30", pady="15", font="15")


button1.pack(padx="15", pady="15")
button2.pack(padx="15", pady="15")


root.mainloop()
